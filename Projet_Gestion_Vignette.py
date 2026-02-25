
       import tkinter as tk
from tkinter import ttk, messagebox, font
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

excel_file = "gestion de vignette.xlsx"

#=================================
#       FICHIER EXCEL
#=================================

def initialize_excel():
    """Crée le fichier Excel s'il n'existe pas"""
    if not os.path.exists(excel_file):
        wb = Workbook()
        wb.remove(wb.active)
    
        sheets = {
            "Proprietaire": ["ID", "Nom", "Commune", "Adresse", "Telephone"],
            "Vehicule": ["ID", "Plaque", "Marque", "Proprietaire_ID"],
            "Vignette": ["ID", "Vehicule_ID", "Annee", "Montant"],
            "Paiement": ["ID", "Vignette_ID", "Montant_Paye", "Date"],
            "Amende": ["ID", "Vehicule_ID", "Jours_Retard", "Montant"] 
        }

        for name, headers in sheets.items():
            ws = wb.create_sheet(name)
            ws.append(headers)
        
        wb.save(excel_file)

def excel_append(sheet, data):
    """Ajoute une ligne au fichier Excel"""
    wb = load_workbook(excel_file)
    ws = wb[sheet]
    ws.append(data)
    wb.save(excel_file)
        
def excel_get(sheet):
    """Récupère les données du fichier Excel"""
    wb = load_workbook(excel_file)
    ws = wb[sheet]
    return list(ws.iter_rows(min_row=2, values_only=True))

def get_next_id(sheet):
    """Retourne le prochain ID disponible"""
    data = excel_get(sheet)
    return data[-1][0] + 1 if data else 1

#=====================================
#       FONCTIONS AJOUTER
#=====================================

def ajouter_proprietaire():
    """Ajoute un propriétaire"""
    nom = entry_nom.get()
    commune = entry_commune.get()
    adresse = entry_adresse.get()
    tel = entry_telephone.get()
    
    if not nom or not commune:
        messagebox.showerror("Erreur", "Nom et Commune requis")
        return
    
    new_id = get_next_id("Proprietaire")
    excel_append("Proprietaire", [new_id, nom, commune, adresse, tel])
    tree_proprietaire.insert("", "end", values=[new_id, nom, commune, adresse, tel])
    
    entry_nom.delete(0, tk.END)
    entry_commune.delete(0, tk.END)
    entry_adresse.delete(0, tk.END)
    entry_telephone.delete(0, tk.END)
    
    messagebox.showinfo("Succès", "Propriétaire ajouté avec succès!")
    refresh_combobox_proprietaire()

def ajouter_vehicule():
    """Ajoute un véhicule"""
    plaque = entry_plaque.get()
    marque = entry_marque.get()
    prop_id_str = combobox_proprietaire.get()
    
    if not plaque or not marque or not prop_id_str:
        messagebox.showerror("Erreur", "Tous les champs requis")
        return
    
    prop_id = prop_id_str.split("-")[0]
    new_id = get_next_id("Vehicule")
    excel_append("Vehicule", [new_id, plaque, marque, prop_id])
    tree_vehicule.insert("", "end", values=[new_id, plaque, marque, prop_id])
    
    entry_plaque.delete(0, tk.END)
    entry_marque.delete(0, tk.END)
    combobox_proprietaire.set("")
    
    messagebox.showinfo("Succès", "Véhicule ajouté avec succès!")
    refresh_combobox_vehicule()

def ajouter_vignette():
    """Ajoute une vignette"""
    vehicule_id_str = combobox_vehicule_vignette.get()
    annee = entry_annee.get()
    montant = entry_montant.get()
    
    if not vehicule_id_str or not annee or not montant:
        messagebox.showerror("Erreur", "Tous les champs requis")
        return
    
    vehicule_id = vehicule_id_str.split("-")[0]
    new_id = get_next_id("Vignette")
    excel_append("Vignette", [new_id, vehicule_id, annee, montant])
    tree_vignette.insert("", "end", values=[new_id, vehicule_id, annee, montant])
    
    entry_annee.delete(0, tk.END)
    entry_montant.delete(0, tk.END)
    combobox_vehicule_vignette.set("")
    
    messagebox.showinfo("Succès", "Vignette ajoutée avec succès!")
    refresh_combobox_vignette()

def valider_paiement():
    """Valide un paiement"""
    vignette_id_str = combobox_vignette.get()
    montant_paye = entry_montant_paye.get()
    
    if not vignette_id_str or not montant_paye:
        messagebox.showerror("Erreur", "Tous les champs requis")
        return
    
    vignette_id = vignette_id_str.split("-")[0]
    new_id = get_next_id("Paiement")
    date_now = datetime.now().strftime("%Y-%m-%d")
    excel_append("Paiement", [new_id, vignette_id, montant_paye, date_now])
    tree_paiement.insert("", "end", values=[new_id, vignette_id, montant_paye, date_now])
    
    entry_montant_paye.delete(0, tk.END)
    combobox_vignette.set("")
    
    messagebox.showinfo("Succès", "Paiement validé avec succès!")

def calculer_amende():
    """Calcule une amende"""
    vehicule_id_str = combobox_vehicule_amende.get()
    jours = entry_jours_retard.get()
    
    if not vehicule_id_str or not jours:
        messagebox.showerror("Erreur", "Tous les champs requis")
        return
    
    try:
        jours_int = int(jours)
    except ValueError:
        messagebox.showerror("Erreur", "Les jours doivent être un nombre")
        return
    
    vehicule_id = vehicule_id_str.split("-")[0]
    montant = jours_int * 10000
    new_id = get_next_id("Amende")
    excel_append("Amende", [new_id, vehicule_id, jours, montant])
    tree_amende.insert("", "end", values=[new_id, vehicule_id, jours, montant])
    
    entry_jours_retard.delete(0, tk.END)
    combobox_vehicule_amende.set("")
    
    messagebox.showinfo("Succès", f"Amende calculée: {montant} FC")

#=========================================
#       RAFRAÎCHISSEMENT COMBOBOX
#=========================================

def refresh_combobox_proprietaire():
    """Rafraîchit la combobox des propriétaires"""
    data = excel_get("Proprietaire")
    combobox_proprietaire['values'] = [f"{d[0]}-{d[1]}" for d in data]

def refresh_combobox_vehicule():
    """Rafraîchit les combobox des véhicules"""
    data = excel_get("Vehicule")
    if 'combobox_vehicule_vignette' in globals():
        combobox_vehicule_vignette['values'] = [f"{d[0]}-{d[1]}" for d in data]
    if 'combobox_vehicule_amende' in globals():
        combobox_vehicule_amende['values'] = [f"{d[0]}-{d[1]}" for d in data]

def refresh_combobox_vignette():
    """Rafraîchit la combobox des vignettes"""
    data = excel_get("Vignette")
    combobox_vignette['values'] = [f"{d[0]}-{d[1]}" for d in data]

#=========================================
#       INTERFACE PRINCIPALE
#=========================================

initialize_excel()

root = tk.Tk()
root.title("Système de Gestion des Vignettes de Kinshasa")
root.geometry("1200x750")
root.configure(bg="#808080")

# Configuration de style
style = ttk.Style()
style.theme_use('clam')

# Définir des couleurs pour chaque section
colors = {
    "Proprietaire": {"bg": "purple", "heading": "#1565C0"},
    "Vehicule": {"bg": "#F3E5F5", "heading": "#6A1B9A"},
    "Vignette": {"bg": "#E8F5E9", "heading": "#00695C"},
    "Paiement": {"bg": "#FFF3E0", "heading": "#E65100"},
    "Amende": {"bg": "#808080", "heading": "#C62828"}
}

style.configure('Title.TLabel', font=('Helvetica', 24, 'bold'), background="#f0f0f0", foreground="#003366")
style.configure('Heading.TLabel', font=('Helvetica', 14, 'bold'), background="#e8f0f8", foreground="#003366")
style.configure('TLabel', font=('Helvetica', 10), background="#f0f0f0")
style.configure('TButton', font=('Helvetica', 10))
style.configure('Treeview', font=('Helvetica', 9), rowheight=25)
style.configure('Treeview.Heading', font=('Helvetica', 10, 'bold'))

# Titre principal
title_frame = tk.Frame(root, bg="#003366", height=80)
title_frame.pack(fill="x", padx=0, pady=0)

title_label = tk.Label(title_frame, text="📋 SYSTÈME DE GESTION DES VIGNETTES - KINSHASA", 
                       font=('Helvetica', 18, 'bold'), bg="#003366", fg="white", pady=20)
title_label.pack()

# Frame pour les onglets
main_frame = tk.Frame(root, bg="#f0f0f0")
main_frame.pack(fill="both", expand=True, padx=10, pady=10)

tabControl = ttk.Notebook(main_frame)
tab = {}

# Créer les onglets
for name in ["Proprietaire", "Vehicule", "Vignette", "Paiement", "Amende"]:
    tab[name] = ttk.Frame(tabControl)
    tabControl.add(tab[name], text=f"  {name}  ")

tabControl.pack(fill="both", expand=True)

#=====================================
#     ONGLET PROPRIETAIRE
#=====================================

frame_prop_form = ttk.LabelFrame(tab["Proprietaire"], text="➕ Ajouter un Propriétaire", padding=15)
frame_prop_form.pack(padx=15, pady=15, fill="x")

# Grille 2 colonnes pour les champs
input_frame = tk.Frame(frame_prop_form, bg=colors["Proprietaire"]["bg"])
input_frame.pack(fill="x", padx=10, pady=10)

tk.Label(input_frame, text="Nom :", font=('Helvetica', 10, 'bold'), bg=colors["Proprietaire"]["bg"], fg=colors["Proprietaire"]["heading"]).grid(row=0, column=0, sticky="w", padx=10, pady=8)
entry_nom = ttk.Entry(input_frame, width=35)
entry_nom.grid(row=0, column=1, padx=10, pady=8)

tk.Label(input_frame, text="Commune :", font=('Helvetica', 10, 'bold'), bg=colors["Proprietaire"]["bg"], fg=colors["Proprietaire"]["heading"]).grid(row=0, column=2, sticky="w", padx=10, pady=8)
entry_commune = ttk.Entry(input_frame, width=35)
entry_commune.grid(row=0, column=3, padx=10, pady=8)

tk.Label(input_frame, text="Adresse :", font=('Helvetica', 10, 'bold'), bg=colors["Proprietaire"]["bg"], fg=colors["Proprietaire"]["heading"]).grid(row=1, column=0, sticky="w", padx=10, pady=8)
entry_adresse = ttk.Entry(input_frame, width=35)
entry_adresse.grid(row=1, column=1, padx=10, pady=8)

tk.Label(input_frame, text="Téléphone :", font=('Helvetica', 10, 'bold'), bg=colors["Proprietaire"]["bg"], fg=colors["Proprietaire"]["heading"]).grid(row=1, column=2, sticky="w", padx=10, pady=8)
entry_telephone = ttk.Entry(input_frame, width=35)
entry_telephone.grid(row=1, column=3, padx=10, pady=8)

button_frame = tk.Frame(frame_prop_form, bg=colors["Proprietaire"]["bg"])
button_frame.pack(fill="x", pady=15)
ttk.Button(button_frame, text="✅ Ajouter Propriétaire", command=ajouter_proprietaire).pack(side="left", padx=10)

# Tableau des propriétaires
frame_prop_table = ttk.LabelFrame(tab["Proprietaire"], text="📊 Liste des Propriétaires", padding=15)
frame_prop_table.pack(padx=15, pady=15, fill="both", expand=True)

tree_proprietaire = ttk.Treeview(frame_prop_table, columns=["ID", "Nom", "Commune", "Adresse", "Telephone"], 
                                 show="headings", height=18)
for col in ["ID", "Nom", "Commune", "Adresse", "Telephone"]:
    tree_proprietaire.heading(col, text=col)
    tree_proprietaire.column(col, width=180)

tree_proprietaire.pack(fill="both", expand=True)

for d in excel_get("Proprietaire"):
    tree_proprietaire.insert("", "end", values=d)

#=====================================
#     ONGLET VEHICULE
#=====================================

frame_veh_form = ttk.LabelFrame(tab["Vehicule"], text="➕ Ajouter un Véhicule", padding=15)
frame_veh_form.pack(padx=15, pady=15, fill="x")

input_frame_veh = tk.Frame(frame_veh_form, bg=colors["Vehicule"]["bg"])
input_frame_veh.pack(fill="x", padx=10, pady=10)

tk.Label(input_frame_veh, text="Plaque :", font=('Helvetica', 10, 'bold'), bg=colors["Vehicule"]["bg"], fg=colors["Vehicule"]["heading"]).grid(row=0, column=0, sticky="w", padx=10, pady=8)
entry_plaque = ttk.Entry(input_frame_veh, width=35)
entry_plaque.grid(row=0, column=1, padx=10, pady=8)

tk.Label(input_frame_veh, text="Marque :", font=('Helvetica', 10, 'bold'), bg=colors["Vehicule"]["bg"], fg=colors["Vehicule"]["heading"]).grid(row=0, column=2, sticky="w", padx=10, pady=8)
entry_marque = ttk.Entry(input_frame_veh, width=35)
entry_marque.grid(row=0, column=3, padx=10, pady=8)

tk.Label(input_frame_veh, text="Propriétaire :", font=('Helvetica', 10, 'bold'), bg=colors["Vehicule"]["bg"], fg=colors["Vehicule"]["heading"]).grid(row=1, column=0, sticky="w", padx=10, pady=8)
combobox_proprietaire = ttk.Combobox(input_frame_veh, width=33)
combobox_proprietaire.grid(row=1, column=1, padx=10, pady=8)
refresh_combobox_proprietaire()

button_frame_veh = tk.Frame(frame_veh_form, bg=colors["Vehicule"]["bg"])
button_frame_veh.pack(fill="x", pady=15)
ttk.Button(button_frame_veh, text="✅ Ajouter Véhicule", command=ajouter_vehicule).pack(side="left", padx=10)

frame_veh_table = ttk.LabelFrame(tab["Vehicule"], text="📊 Liste des Véhicules", padding=15)
frame_veh_table.pack(padx=15, pady=15, fill="both", expand=True)

tree_vehicule = ttk.Treeview(frame_veh_table, columns=["ID", "Plaque", "Marque", "Proprietaire_ID"], 
                             show="headings", height=18)
for col in ["ID", "Plaque", "Marque", "Proprietaire_ID"]:
    tree_vehicule.heading(col, text=col)
    tree_vehicule.column(col, width=280)

tree_vehicule.pack(fill="both", expand=True)

for d in excel_get("Vehicule"):
    tree_vehicule.insert("", "end", values=d)

#=====================================
#     ONGLET VIGNETTE
#=====================================

frame_vig_form = ttk.LabelFrame(tab["Vignette"], text="➕ Ajouter une Vignette", padding=15)
frame_vig_form.pack(padx=15, pady=15, fill="x")

input_frame_vig = tk.Frame(frame_vig_form, bg=colors["Vignette"]["bg"])
input_frame_vig.pack(fill="x", padx=10, pady=10)

tk.Label(input_frame_vig, text="Véhicule :", font=('Helvetica', 10, 'bold'), bg=colors["Vignette"]["bg"], fg=colors["Vignette"]["heading"]).grid(row=0, column=0, sticky="w", padx=10, pady=8)
combobox_vehicule_vignette = ttk.Combobox(input_frame_vig, width=33)
combobox_vehicule_vignette.grid(row=0, column=1, padx=10, pady=8)
refresh_combobox_vehicule()

tk.Label(input_frame_vig, text="Année :", font=('Helvetica', 10, 'bold'), bg=colors["Vignette"]["bg"], fg=colors["Vignette"]["heading"]).grid(row=0, column=2, sticky="w", padx=10, pady=8)
entry_annee = ttk.Entry(input_frame_vig, width=35)
entry_annee.grid(row=0, column=3, padx=10, pady=8)

tk.Label(input_frame_vig, text="Montant :", font=('Helvetica', 10, 'bold'), bg=colors["Vignette"]["bg"], fg=colors["Vignette"]["heading"]).grid(row=1, column=0, sticky="w", padx=10, pady=8)
entry_montant = ttk.Entry(input_frame_vig, width=35)
entry_montant.grid(row=1, column=1, padx=10, pady=8)

button_frame_vig = tk.Frame(frame_vig_form, bg=colors["Vignette"]["bg"])
button_frame_vig.pack(fill="x", pady=15)
ttk.Button(button_frame_vig, text="✅ Ajouter Vignette", command=ajouter_vignette).pack(side="left", padx=10)

frame_vig_table = ttk.LabelFrame(tab["Vignette"], text="📊 Liste des Vignettes", padding=15)
frame_vig_table.pack(padx=15, pady=15, fill="both", expand=True)

tree_vignette = ttk.Treeview(frame_vig_table, columns=["ID", "Vehicule_ID", "Annee", "Montant"], 
                             show="headings", height=18)
for col in ["ID", "Vehicule_ID", "Annee", "Montant"]:
    tree_vignette.heading(col, text=col)
    tree_vignette.column(col, width=280)

tree_vignette.pack(fill="both", expand=True)

for d in excel_get("Vignette"):
    tree_vignette.insert("", "end", values=d)

#================================
#     ONGLET PAIEMENT
#================================

frame_paye_form = ttk.LabelFrame(tab["Paiement"], text="➕ Enregistrer un Paiement", padding=15)
frame_paye_form.pack(padx=15, pady=15, fill="x")

input_frame_paye = tk.Frame(frame_paye_form, bg=colors["Paiement"]["bg"])
input_frame_paye.pack(fill="x", padx=10, pady=10)

tk.Label(input_frame_paye, text="Vignette :", font=('Helvetica', 10, 'bold'), bg=colors["Paiement"]["bg"], fg=colors["Paiement"]["heading"]).grid(row=0, column=0, sticky="w", padx=10, pady=8)
combobox_vignette = ttk.Combobox(input_frame_paye, width=33)
combobox_vignette.grid(row=0, column=1, padx=10, pady=8)
refresh_combobox_vignette()

tk.Label(input_frame_paye, text="Montant payé :", font=('Helvetica', 10, 'bold'), bg=colors["Paiement"]["bg"], fg=colors["Paiement"]["heading"]).grid(row=0, column=2, sticky="w", padx=10, pady=8)
entry_montant_paye = ttk.Entry(input_frame_paye, width=35)
entry_montant_paye.grid(row=0, column=3, padx=10, pady=8)

button_frame_paye = tk.Frame(frame_paye_form, bg=colors["Paiement"]["bg"])
button_frame_paye.pack(fill="x", pady=15)
ttk.Button(button_frame_paye, text="✅ Valider Paiement", command=valider_paiement).pack(side="left", padx=10)

frame_paye_table = ttk.LabelFrame(tab["Paiement"], text="📊 Historique des Paiements", padding=15)
frame_paye_table.pack(padx=15, pady=15, fill="both", expand=True)

tree_paiement = ttk.Treeview(frame_paye_table, columns=["ID", "Vignette_ID", "Montant_Paye", "Date"], 
                             show="headings", height=18)
for col in ["ID", "Vignette_ID", "Montant_Paye", "Date"]:
    tree_paiement.heading(col, text=col)
    tree_paiement.column(col, width=280)

tree_paiement.pack(fill="both", expand=True)

for d in excel_get("Paiement"):
    tree_paiement.insert("", "end", values=d)

#=================================
#     ONGLET AMENDE
#=================================

frame_amend_form = ttk.LabelFrame(tab["Amende"], text="➕ Calculer une Amende", padding=15)
frame_amend_form.pack(padx=15, pady=15, fill="x")

input_frame_amend = tk.Frame(frame_amend_form, bg=colors["Amende"]["bg"])
input_frame_amend.pack(fill="x", padx=10, pady=10)

tk.Label(input_frame_amend, text="Véhicule :", font=('Helvetica', 10, 'bold'), bg=colors["Amende"]["bg"], fg=colors["Amende"]["heading"]).grid(row=0, column=0, sticky="w", padx=10, pady=8)
combobox_vehicule_amende = ttk.Combobox(input_frame_amend, width=33)
combobox_vehicule_amende.grid(row=0, column=1, padx=10, pady=8)
refresh_combobox_vehicule()

tk.Label(input_frame_amend, text="Jours de Retard :", font=('Helvetica', 10, 'bold'), bg=colors["Amende"]["bg"], fg=colors["Amende"]["heading"]).grid(row=0, column=2, sticky="w", padx=10, pady=8)
entry_jours_retard = ttk.Entry(input_frame_amend, width=35)
entry_jours_retard.grid(row=0, column=3, padx=10, pady=8)

button_frame_amend = tk.Frame(frame_amend_form, bg=colors["Amende"]["bg"])
button_frame_amend.pack(fill="x", pady=15)
ttk.Button(button_frame_amend, text="✅ Calculer Amende", command=calculer_amende).pack(side="left", padx=10)

frame_amend_table = ttk.LabelFrame(tab["Amende"], text="📊 Registre des Amendes", padding=15)
frame_amend_table.pack(padx=15, pady=15, fill="both", expand=True)

tree_amende = ttk.Treeview(frame_amend_table, columns=["ID", "Vehicule_ID", "Jours_Retard", "Montant"], 
                           show="headings", height=18)
for col in ["ID", "Vehicule_ID", "Jours_Retard", "Montant"]:
    tree_amende.heading(col, text=col)
    tree_amende.column(col, width=280)

tree_amende.pack(fill="both", expand=True)

for d in excel_get("Amende"):
    tree_amende.insert("", "end", values=d)

#=========================================
#       LANCER L'APPLICATION
#=========================================

root.mainloop()            
